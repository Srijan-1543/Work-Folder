from openpyxl import load_workbook
#reading
wb=load_workbook(r'D:\Software_Coding Files\Verzeo-Python with machine learning\dumtest.xlsx')
ws=wb["Sheet1"]
print(ws.cell(3,3).value)
print(type(ws.cell(3,3)))
print(type(wb))

#writing
ws.cell(5,3,"dogbee")
wb.save(r'D:\Software_Coding Files\Verzeo-Python with machine learning\dumtest.xlsx')
